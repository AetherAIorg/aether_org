from __future__ import annotations

import io
import re

from openpyxl import load_workbook

from app.parsers import RawImplementation


def _infer_label(ws, row: int, col: int) -> str:
    for c in range(col - 1, 0, -1):
        value = ws.cell(row=row, column=c).value
        if value and isinstance(value, str) and not value.startswith("="):
            return value.strip()
    for r in range(row - 1, 0, -1):
        value = ws.cell(row=r, column=col).value
        if value and isinstance(value, str) and not value.startswith("="):
            return value.strip()
    return f"Formula_{ws.title}!{row}_{col}"


def _extract_sheet_refs(formula: str) -> list[str]:
    refs = re.findall(r"(?:'[^']+'|[A-Za-z_][\w]*)!", formula)
    return sorted({ref.rstrip("!") for ref in refs})


def _extract_functions(formula: str) -> list[str]:
    return sorted(set(re.findall(r"\b([A-Z][A-Z0-9_]*)\s*\(", formula)))


def parse_excel(content: bytes) -> list[RawImplementation]:
    wb = load_workbook(io.BytesIO(content), data_only=False)
    results: list[RawImplementation] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
                    label = _infer_label(ws, cell.row, cell.column)
                    location = f"{sheet_name}!{cell.coordinate}"
                    deprecated = bool(re.search(r"\b(legacy|deprecated|old)\b", label, re.I))
                    results.append(
                        RawImplementation(
                            location=location,
                            label=label,
                            raw_formula=formula,
                            source_refs=_extract_sheet_refs(formula),
                            referenced_functions=_extract_functions(formula),
                            is_deprecated=deprecated,
                        )
                    )
    return results
