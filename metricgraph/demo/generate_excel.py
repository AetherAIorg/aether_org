"""Generate Fund_Performance_Model.xlsx demo artifact."""
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Summary"

labels = [
    ("Fund Name", "A"),
    ("Net IRR", "B"),
    ("Gross IRR", "C"),
    ("MOIC", "D"),
    ("DPI", "E"),
    ("TVPI", "F"),
    ("NAV", "G"),
]
for i, (label, _) in enumerate(labels, start=1):
    ws.cell(row=i, column=1, value=label)

ws["H21"] = "Net IRR"
ws["H22"] = "=XIRR(NetCashflows, CashflowDates)"
ws["H23"] = "Gross Deal IRR"
ws["H24"] = "=XIRR(GrossCashflows, CashflowDates)"
ws["H25"] = "Realized IRR"
ws["H26"] = "=XIRR(RealizedCashflows, ExitDates)"
ws["H27"] = "Fund MOIC"
ws["H28"] = "=(SUM(Distributions)+NAV)/PaidInCapital"
ws["H29"] = "DPI"
ws["H30"] = "=SUM(Distributions)/PaidInCapital"
ws["H31"] = "TVPI"
ws["H32"] = "=(SUM(Distributions)+NAV)/PaidInCapital"
ws["H33"] = "Legacy Monthly IRR"
ws["H34"] = "=MIRR(MonthlyCashflows, FinanceRate, ReinvestRate)"

ws2 = wb.create_sheet("Cashflows")
ws2["A1"] = "Date"
ws2["B1"] = "Net Amount"
ws2["C1"] = "Gross Amount"
ws2["A2"] = "2020-01-15"
ws2["B2"] = -10000000
ws2["C2"] = -9500000

wb.save("demo/investment_ops_demo/Fund_Performance_Model.xlsx")
print("Created Fund_Performance_Model.xlsx")

if __name__ == "__main__":
    pass
